"""
loader.py - загружает данные из excel-файла в список словарей.
"""
from openpyxl import load_workbook


def load_data_from_excel(
        input_path: str,
        sheet_name: str,
        header_row: int,
        min_col: int,
        max_col: int
) -> list[dict]:
    """Загружает данные из excel-файла в список словарей.

    input_path: путь к excel-файлу с таблицей,
    sheet_name: имя листа, где располагается искомая тиблица,
    header_row: номер строки с заголовками (нумерация с 1),
    min_col: номер столбца начала таблицы (нумерация с 1),
    max_col: номер столбца окончания таблицы

    Возвращает список словарей, где каждый словарь - строка из
    исходной таблицы.
    """
    wb = load_workbook(input_path, data_only=True)  # загрузить только значения

    ws = wb[sheet_name]  # открыть указанный лист

    headers_row = next(ws.iter_rows(
            min_row=header_row,
            max_row=header_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True
        ))
    # сформировать список заголовков, чтобы заголовки совпадали по ширине с остальной таблицей

    headers = list(headers_row)

    table = []

    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=ws.max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True
        ):

        row_dict = dict(zip(headers, row))  # связать заголовки со значением, превратить в словарь

        table.append(row_dict)  # формируем таблицу, как список строк-словарей

    return table
