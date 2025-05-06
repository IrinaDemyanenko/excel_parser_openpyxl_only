"""
preprocessor.py — предварительная обработка excel-файла перед загрузкой.

Цель: очистить колонку excel от лишних текстовых символов (например, '600op.h.'),
чтобы excel мог корректно пересчитать формулы, основанные на этих данных.

Результат сохраняется в новый файл (копию исходного),
который затем используется для дальнейшей обработки.
"""

from openpyxl import load_workbook
import re
from openpyxl.workbook.properties import CalcProperties


INT_PATTERN = r'([0-9]+)'
FLOAT_PATTERN = r'([0-9]*\.?[0-9]+)'

def clean_column_and_save_copy(
        input_path: str,
        output_path: str = None,
        sheet_name: str = 'Sheet1',
        column_letter: str = 'A',
        start_row: int = 1,
        pattern: str = INT_PATTERN
) -> str:
    """Очищает колонку в файле excel от текста, оставляя только число.

    После очистки колонки результат сохраняется в новый excel файл.

    input_path: путь к исходному файлу
    output_path: путь к очищенному файлу
    sheet_name: имя листа, по умолчанию первый лист
    column_letter: буква колонки (столбца), по умолчанию ''
    start_row: строка, с которой начинаются данные
    pattern: регулярное выражение (что извлекать из данных?)
    Возвращает путь к новому очищенному файлу.
    """

    # если путь вообще не задан, формируем его
    if output_path is None:
        output_path = input_path.replace('.xlsx', '_cleaned.xlsx')

    wb = load_workbook(input_path)  # объект Excel-файла, созданный через openpyxl
    ws = wb[sheet_name]  # открываем лист с нужной таблицей

    for row in range(start_row, ws.max_row + 1):
        cell = ws[f'{column_letter}{row}']  # ячейка в нужном столбце
        value = str(cell.value)  # значение в ячейке -> в строку
        match = re.search(pattern, value)  # найди совпадение в строке по шаблону

        # условия и для поиска по шаблону float_pattern
        if match:
            cell.value = float(match.group(1)) if '.' in match.group(1) else int(match.group(1))   # group(0) — это всегда всё совпадение, group(1) и дальше — это то, что в скобки поймала
        else:
            cell.value = None

    wb.calc_properties = CalcProperties(fullCalcOnLoad=True) # Указываем Excel: нужно пересчитать формулы

    wb.save(output_path)  # сохраняем файл с очищенной колонкой

    print(f'Новый файл-excel сохранён здесь: {output_path}')

    return output_path
